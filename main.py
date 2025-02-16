import pandas as pd
import numpy as np
import openpyxl
'''Блок импорта'''
# Непустые столбцы
NEEDED_COLUMNS = [0, 1, 2, 3, 4, 6, 7]
# Импорт из файла
money_income_df = pd.read_excel('data.xlsx', usecols=NEEDED_COLUMNS)
''''''
'''Блок добавления месяца сделки и удаления строк с ними'''
# Конвертация даты формата Месяц Год в формат datetime
def convert_date(date_str):
    # Словарь месяцев ru-en
    months = {
        'Январь': 'January', 'Февраль': 'February', 'Март': 'March', 'Апрель': 'April',
        'Май': 'May', 'Июнь': 'June', 'Июль': 'July', 'Август': 'August',
        'Сентябрь': 'September', 'Октябрь': 'October', 'Ноябрь': 'November', 'Декабрь': 'December'
    }
    # Кортеж месяца и года из строки
    month, year = date_str.split()
    # Приведение к типу datetime
    return pd.to_datetime(f"{months[month]} {year}", format='%B %Y')
# Строки с месяцами
months_with_ids_df = money_income_df[money_income_df['client_id'].isna()][['status']]
# Строки с месяцами в формате datetime
months_with_ids_df['status'] = months_with_ids_df['status'].apply(convert_date)
# Кол-во месяцев
num_months = len(months_with_ids_df)
# Для каждого месяца
for i in range(num_months):
    # Находим его индекс
    i_index = months_with_ids_df.index[i]
    # Если не последний
    if i != num_months - 1:
        # От индекса текущего до индекса следующего присваиваем текущий месяц
        money_income_df.loc[i_index:months_with_ids_df.index[i+1], 'month'] = months_with_ids_df.loc[i_index, 'status']
    # Если последний
    else:
        # От индекса текущего месяца присваеваем до конца дфа
        money_income_df.loc[i_index:, 'month'] = months_with_ids_df.loc[i_index, 'status']
# Убираем строки с месяцами
money_income_df = money_income_df.dropna(subset=['client_id'])
# Приведение даты в столбце к формату datetime для удобства сравнения дат
money_income_df['receiving_date'] = pd.to_datetime(money_income_df['receiving_date'], errors='coerce')
''''''
'''Блок действий для 1 вопроса'''
# Непросроченные поступления
not_expired_df = money_income_df[money_income_df['status'] != 'ПРОСРОЧЕНО']
# 1) Вычислите общую выручку за июль 2021 по тем сделкам, приход денежных средств, которых не просрочен
question_1 = not_expired_df.loc[(not_expired_df['month'] == '2021-07-01'), 'sum'].sum()
''''''
'''Блок действий для 2 вопроса'''
# 2) Как изменялась выручка компании за рассматриваемый период? Проиллюстрируйте графиком.
question_2 = money_income_df.groupby('month', as_index=False)['sum'].sum()
question_2['sum_thousand'] = question_2['sum'] / 1000
print(question_2)
''''''
'''Блок действий для 3 вопроса'''
# Фильтрация по Сентябрю 2021
question_3 = money_income_df.loc[(money_income_df['month'] == '2021-09-01'), ['sale', 'sum']]
# 3) Кто из менеджеров привлек для компании больше всего денежных средств в сентябре 2021?
question_3_1 = question_3.groupby('sale').sum()
print(question_3_1['sum'].idxmax())
''''''
'''Блок действий для 4 вопроса'''
# Фильтрация по Октябрю 2021
question_4_1 = money_income_df.loc[money_income_df['month'] == '2021-07-01'].groupby('new/current').count()
# 4) Какой тип сделок (новая/текущая) был преобладающим в октябре 2021?
print(question_4_1['month'].idxmax())
''''''
'''Блок действий для 5 вопроса'''
# Фильтрация по Маю 2021 и по дате получения оригинала в июне 2021
question_5 = money_income_df.loc[(money_income_df['month'] == '2021-05-01') & \
                                 (money_income_df['receiving_date'].dt.year == 2021) & \
                                 (money_income_df['receiving_date'].dt.month == 6)]
# 5) Сколько оригиналов договора по майским сделкам было получено в июне 2021?
print(len(question_5))
''''''
'''Блок действий по заданию'''
# Копия дф с непустыми менеджерами
money_without_1 = money_income_df.loc[money_income_df['sale'] != '-'].copy()
# Условия для бонусов
# Оригинал пришел в месяц заключения новой оплаченной сделки
cond1 = (money_without_1['status'] == 'ОПЛАЧЕНО') & \
        (money_without_1['receiving_date'].dt.year == money_without_1['month'].dt.year) & \
        (money_without_1['receiving_date'].dt.month == money_without_1['month'].dt.month) & \
        (money_without_1['new/current'] == 'новая')
# Непросроченная текущая сделка с суммой <= 10000
cond2 = (money_without_1['status'] != 'ПРОСРОЧЕНО') & \
        (money_without_1['new/current'] == 'текущая') & \
        (money_without_1['sum'] > 10_000)
# Непросроченная текущая сделка с суммой > 10000
cond3 = (money_without_1['status'] != 'ПРОСРОЧЕНО') & \
        (money_without_1['new/current'] == 'текущая') & \
        (money_without_1['sum'] <= 10_000)
# Выдаем бонусы в соответствии с условием
money_without_1['bonus'] = np.select(
    [cond1, cond2, cond3],
    [
     0.07 * money_without_1['sum'],  # 7% от суммы для 1 условия
     0.05 * money_without_1['sum'],  # 5% от суммы для 2 условия
     0.03 * money_without_1['sum']   # 3% от суммы для 3 условия
    ],
    default=0  # Иначе бонус равен 0
)
# Получаем список менеджеров
unique_sales = money_without_1['sale'].unique()
# Группируем по менеджерам и считаем сумму их бонусов к дате 2021-07-01 включительно
grouped_sum = money_without_1.loc[money_without_1['receiving_date'] <= '2021-07-01'].groupby('sale')['sum'].sum().reset_index()
# Создаем дф с менеджерами
result = pd.DataFrame(unique_sales, columns=['sale'])
# Присваиваем сумму или 0
result['sum'] = result['sale'].map(grouped_sum.set_index('sale')['sum']).fillna(0)
print(result)
'''Блок записи ответов в xlsx файл'''
# Создаем новый файл
wb = openpyxl.Workbook()
'''Блок записи 1 ответа'''
# Выбираем активный лист
sheet = wb.active
# Даем название листу
sheet.title = "Вопрос 1"
# Записываем данные
sheet["A1"] = "Сумма за июль 2021 с статусом не ПРОСРОЧЕНО"
sheet["A2"] = question_1
''''''
'''Блок записи 2 ответа'''
# Создаем новый лист
sheet = wb.create_sheet("Вопрос 2")

# Записываем заголовки
sheet.cell(row=1, column=1, value="Месяц")
sheet.cell(row=1, column=2, value="Сумма")

# Записываем данные
for row_idx, (month, total) in enumerate(zip(question_2['month'], question_2['sum']), start=2):
    sheet.cell(row=row_idx, column=1, value=month)
    sheet.cell(row=row_idx, column=2, value=total)

# Создаем диаграмму
chart = openpyxl.chart.LineChart()
chart.title = "Доходы по месяцам"
chart.y_axis.title = "Сумма в тыс. рублей"
chart.x_axis.title = "Месяцы"

# Указываем диапазон данных для диаграммы
data = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_row=len(question_2) + 1, max_col=2)
categories = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=len(question_2) + 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Добавляем диаграмму на лист
sheet.add_chart(chart, "D4")
''''''
'''Блок записи 3 ответа'''
# Создаем новый лист
sheet = wb.create_sheet("Вопрос 3")
# Записываем данные
sheet["A1"] = "Менеджер, который привлек больше средств с любым статусом в сентябре 2021"
sheet["A2"] = question_3_1['sum'].idxmax()
''''''
'''Блок записи 4 ответа'''
# Создаем новый лист
sheet = wb.create_sheet("Вопрос 4")
# Записываем данные
sheet["A1"] = "Преобладающий тип сделки в октябре 2021"
sheet["A2"] = question_4_1['month'].idxmax()
''''''
'''Блок записи 5 ответа'''
# Создаем новый лист
sheet = wb.create_sheet("Вопрос 5")
# Записываем данные
sheet["A1"] = "Кол-во оригиналов договора по майским сделкам было получено в июне 2021"
sheet["A2"] = len(question_5)
''''''
'''Блок записи Ответа на Задание'''
# Создаем новый лист
sheet = wb.create_sheet("Бонусы сотрудников к 2021-07-01")
# Сохраняем файл
wb.save("output.xlsx")
# Закрываем книгу
wb.close()
# Записываем данные в созданный лист
with pd.ExcelWriter("output.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    result.to_excel(writer, sheet_name="Бонусы сотрудников к 2021-07-01", index=False)
''''''
''''''